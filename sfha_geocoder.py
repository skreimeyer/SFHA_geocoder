import sys
import re
from dataclasses import dataclass
from enum import Enum
from typing import List

import openpyxl as xl
import requests


@dataclass
class Point:
    x: float
    y: float


@dataclass
class Legal:
    lot: int
    block: int
    subdivision: str


class AddressType(Enum):
    ADD = 1
    LEGAL = 2
    NA = 3


PATTERN_ADDRESS = re.compile(
    r"^\d+ [\D]+ (?:Road|Drive|Court|Cove|Boulevard|Street|Circle|\d{1,3})$"
)
PATTERN_LEGAL = re.compile(r"((Lot|Block) (\d+) )((Lot|Block) (\d+) ){0,1}(\D+)$")


def run(source_filename: str) -> ():
    """
    core function, loads workbook and drives workbook iteration and relies on
    helper functions for network-related tasks and string parsing
    """
    try:
        wb = xl.load_workbook(source_filename)
    except:
        print(f"Cannot load workbook named: {source_filename}")
        return
    ws = wb.active
    # Ready a destination spreadsheet
    new_wb = xl.Workbook()
    new_ws = new_wb.active
    # warm up a Requests session
    sess = requests.session()
    # Iterate over values in source spreadsheet
    rows = ws.iter_rows(values_only=True)
    headers = next(rows)
    # index address header and count columns
    if 'Address' not in headers: # We likely do not have a usable spreadsheet
        print("The spreadsheet has no column named 'Address'")
        return
    col_add = headers.index("Address")
    if "X" not in headers:
        headers = (*headers,"X")
    if "Y" not in headers:
        headers = (*headers,"Y")
    col_x = headers.index("X")
    col_y = headers.index("Y")
    new_ws.append(headers)
    # main loop
    for r in rows:
        new_row = list(r)
        if len(r) < len(headers):
            new_row.append(None)
            new_row.append(None)
        address_string = r[col_add]
        add_type = classify(address_string)
        if add_type == AddressType.NA:
            new_ws.append(new_row)
            continue
        elif add_type == AddressType.LEGAL:
            legal = parse_legal(address_string)
            point = geocode_legal(sess, legal)
        else:
            point = geocode(sess, address_string)
        if point.x == 0:
            new_ws.append(new_row) # almost miss goto
            continue
        new_row[col_x] = point.x
        new_row[col_y] = point.y
        new_ws.append(new_row)
    dst_name = source_filename.split(".")[0]+"_geocoded.xlsx"
    new_wb.save(filename=dst_name)
    print("Worksheet saved")


def geocode(s: requests.Session, address: str) -> Point:
    """
    queries the PAGIS LOCATOR server with an address string. Returns first
    result.
    """
    url = "http://pagis.org/arcgis/rest/services/LOCATORS/AddressPoints/GeocodeServer/findAddressCandidates"
    query = {
        "category": "",
        "distance": "",
        "location": "",
        "magicKey": "",
        "maxLocations": "",
        "outFIelds": "",
        "outSR": "",
        "searchExtent": "",
        "Single Line Input": address,
        "Street": "",
        "ZIP": "",
        "f": "pjson",
    }
    try:
        resp = s.get(url, params=query)
        j = resp.json()
        if len(j["candidates"]) == 0:
            print(f"No candidates for {address}")
            return Point(0,0)
        loc_dict = j["candidates"][0]["location"]
        location = Point(loc_dict["x"], loc_dict["y"])  # could just return dict
    except Exception as e:  # a famous antipattern
        print(f"Failed to fetch {address} with error:")
        print(e)
        location = Point(0, 0)
    return location


def geocode_legal(s: requests.Session, legal: Legal) -> Point:
    """
    queries the PAGIS Parcels Map Server (51) with a boundary containing the
    entire city where lot, block and subdivision names match
    """
    url = "http://pagis.org/arcgis/rest/services/APPS/OperationalLayers/MapServer/51/query"
    where_clause = f"SUB_NAME LIKE '{legal.subdivision.upper()}%'"
    if legal.lot is not None:
        where_clause += f" AND LOT LIKE '{legal.lot}'"
    if legal.block is not None:
        where_clause += f" AND BLOCK LIKE '{legal.block}'"
    query = {
        "where": where_clause,
        "text": "",
        "objectIds": "",
        "time": "",
        "geometry": """{"xmin":"1150000","ymin":"100000","xmax":"1275000","ymax":"180000","spatialReference":{"wkid":102651,"latestWkid":3433}}""",
        "geometryType": "esriGeometryEnvelope",
        "inSR": "",
        "spatialRel": "esriSpatialRelIntersects",
        "relationParam": "",
        "outFields": "",
        "returnGeometry": True,
        "returnTrueCurves": False,
        "maxAllowableOffset": "",
        "geometryPrecision": "",
        "outSR": "",
        "returnIdsOnly": False,
        "returnCountOnly": False,
        "orderByFields": "",
        "groupByFieldsForStatistics": "",
        "outStatistics": "",
        "returnZ": False,
        "returnM": False,
        "gdbVersion": "",
        "returnDistinctValues": False,
        "resultOffset": "",
        "resultRecordCount": "",
        "queryByDistance": "",
        "returnExtentsOnly": False,
        "datumTransformation": "",
        "parameterValues": "",
        "rangeValues": "",
        "f": "pjson"
    }
    try:
        resp = s.get(url, params=query)
        j = resp.json()
        if len(j["features"]) == 0:
            print(f"No candidates for {legal}")
            return Point(0,0)
        ring = j["features"][0]["geometry"]["rings"]
        point = centroid(ring)
    except Exception as e:
        print(f"failed to find legal: {legal} with error:", e)
        point = Point(0, 0)
    return point


def centroid(ring: List[List[List[float]]]) -> Point:
    """
    average nested x and y pairs to calculate geometry center. Takes an esri
    ring type. Geometries with multiple rings are ignored.
    """
    r = ring[0]
    x_avg = sum([point[0] for point in r]) / len(r)
    y_avg = sum([point[1] for point in r]) / len(r)
    return Point(x_avg, y_avg)


def classify(addr_string: str) -> AddressType:
    """
    uses regex to determine if address string is standard, lot/block or invalid

    Legal descriptions will run the regex twice, but this implementation is
    expected to be more coherent than conditionally passing a regex match.
    """
    if PATTERN_ADDRESS.match(addr_string):
        return AddressType.ADD
    if PATTERN_LEGAL.match(addr_string):
        return AddressType.LEGAL
    return AddressType.NA


def parse_legal(addr_string: str) -> Legal:
    """
    Extracts lot/block numbers and subdivision name
    """
    m = PATTERN_LEGAL.match(addr_string)
    lot, block = None, None  # either lot or block might not be in addrstring
    if m.group(2) == "Lot":
        lot = int(m.group(3))
    else:
        block = int(m.group(3))
    if m.group(4) is not None:
        if m.group(5) == "Lot":
            lot = int(m.group(6))
        else:
            block = int(m.group(6))
    subdivision = m.group(7)
    return Legal(lot, block, subdivision)


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print(
            """
Usage: python sfha_geocoder FILENAME

Takes an excel file containing addresses and attempts to geocode all rows with 
the PAGIS geocoding service. Addresses must have a valid pattern to be coded:

    [Number] [Street Name] ... [Suffix]

or:

    Lot [Number] { Block [Number] } [Subdivision Name] ...
    Block [Number] [Subdivision Name] ...

Rows prefixed with "Lot" or "Block" will result in a query to the parcel server.
The address string will be parsed into lot and block numbers and subdivision
name. The first result will be assumed valid, and the parcel will be geocoded
by calculating the parcel centroid.

Address fields which match neither pattern will be skipped. All data will be
copied to a new spreadsheet named ORIGINALFILE_geocoded.xls
        """
        )
    run(sys.argv[1])
